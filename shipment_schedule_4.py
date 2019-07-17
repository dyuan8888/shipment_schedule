# -*- coding: utf-8 -*-
"""
Updated on April 23, 2019

What's new:
1. Create a ship data dictinonary with JSON
2. Added a function to check the cancelled items


@author: DanielYuan
"""

#!shipment_schedule3.py
"""Extract the data from Product Plan and write them to a new spreadsheet"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import RED, BLUE
from openpyxl.comments import Comment
import datetime
import json


# Open the source file to read data from
def open_src_file(src_file, *sheets):
    wb = load_workbook(src_file, read_only=True, data_only=True)
    print('Start reading data from the source file...')
    for sheet in sheets:
        ws = wb[sheet]
        if sheet == 'M-Plan':
            mo_list = extract_data(ws,rx=3,ry=8)
        else:
            etch_list = extract_data(ws,rx=5,ry=10)    
    sheets_value = mo_list + etch_list
    print('Reading data done!\n')
    return sheets_value


# Extract data from the MFG product plan
def extract_data(ws,rx,ry):
    total_cell_values=[]
    now = datetime.datetime.now()
    future= now +datetime.timedelta(days=260)
    for i in range(2,290):
        cell_values=[]
        for j in range(rx, ry):
            cell_value = ws.cell(row=i, column=j).value
            cell_values.append(cell_value)
        if cell_values[1] != None and 'NSO' not in cell_values[1] and 'Upgr' not in cell_values[1] and now<cell_values[4]<future and 'R&D' not in cell_values[0]:
            cell_values = format_date(cell_values)
            total_cell_values.append(cell_values)
    return total_cell_values


# Format the date
def format_date(cell_values):
    cell_values[3] = datetime.datetime.strftime(cell_values[3],'%Y-%m-%d')
    cell_values[4] = datetime.datetime.strftime(cell_values[4],'%Y-%m-%d')
    return cell_values


# Write data in the target file
def write_data(sheets_value, new_tar_file, compare_file):
    wb = Workbook()
    ws = wb.active
    row_title_style(wb,ws)  # Call to generate column title
    if compare_file:  
        # compare_file is not empty
        print('Starting writing data to the target file...')
        num = 0    
        for row_value in sheets_value:
            row_value = compare_data(row_value, compare_file)
            # TODO compare_SSC_data()
            ws.append(row_value)
            num += 1
            if num < 2:
                print('%s item created...' %num)
            else:
                print('%s items created...' %num)
            color_font_comment(ws, compare_file)
        
    else:
        print('Starting writing data to the target file...')
        num = 0    
        for row_value in sheets_value:
            ws.append(row_value)
            num += 1
            if num < 2:
                print('%s row added...' %num)
            else:
                print('%s rows added...' %num)
    # check any cancelled items
    cancelled_items(sheets_value, compare_file) 
    wb.save(new_tar_file)
    print('\nA total of %d items has been successfully created!' %num)


# Color the specified fonts and add comments to changed schedules
def color_font_comment(ws, compare_file):
    for cell in list(ws.columns)[2]:
        if '*' in cell.value:
            cell.font = Font(color=BLUE)
        elif '!' in cell.value:
            cell.font = Font(color=RED)
            cmt_value1 = compare_file[cell.value.strip('!')][0]
            cmt_value2 = compare_file[cell.value.strip('!')][1]
            cell.comment = Comment('Last crate date:\n %s\nLast ship date:\n%s' % (cmt_value1, cmt_value2), 'Author')


# Compare new ship data with existing ship data
def compare_data(row_value, compare_file):
    row_value[2] = str(row_value[2])
    if row_value[2] in compare_file.keys():
        if row_value[3] != compare_file[row_value[2]][0] or row_value[4] != compare_file[row_value[2]][1]:
            row_value[2] = row_value[2] + '!'
    elif row_value[2] not in compare_file.keys():
        row_value[2] = row_value[2] + '*'
    return row_value

def cancelled_items(sheets_value, compare_file):
    sheets_list = []
    for k in sheets_value:
        k[2] = str(k[2]).strip('*').strip('!')
        sheets_list.append(k[2])
    print('sheets_list',sheets_list)
    for j in compare_file.keys():
        if j not in sheets_list:
            print('%s was cancelled...' %j)


# Style the first-row title
def row_title_style(wb, ws):
    title = ['Customer', 'Product Info', 'Project ID', 'Crate Date', 'Ship Date']
    ws.append(title)
    for col in range(1,6):
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.cell(row =1, column = col).font = Font(bold=True, size=14)


def file_names():
    '''Get the MFG Product Plan source file and 
    the existing Shipment Schedule file for comparison'''
    prompt = ('Please enter the MFG Product Plan source file.\n')
    filename= input(prompt + '>> ')
    src_file = 'c:\\users\\danielyuan\\desktop\\ProductPlanProject\\%s.xlsx' %filename
    now = datetime.datetime.now()
    now_str = datetime.datetime.strftime(now, '%m%d%Y')
    new_tar_file = 'c:\\users\\danielyuan\\desktop\\ProductPlanProject\\ShipmentSchedule_%s.xlsx' % now_str
    return src_file, new_tar_file


def read_cmp_file():
    compare_file = 'compare_data.json'
    try:
        with open(compare_file, 'r') as f_obj:
            compare_file = json.load(f_obj)
    except FileNotFoundError:
        return None
    else:
        return compare_file


def save_cmp_file(compare_data):  
    # Save compare_data into compare_data.json(compare_file)
    compare_file = 'compare_data.json'
    with open(compare_file, 'w') as f_obj:
        json.dump(compare_data, f_obj)


def make_cmp_data(sheets_value):
    cmp_data ={}
    for i in sheets_value:
        i[2] = i[2].strip('*').strip('!')
        cmp_data[i[2]]=[i[3],i[4]]
    return cmp_data


def main():
    src_file, new_tar_file = file_names()
    sheets_value = open_src_file(src_file, 'M-Plan', 'E-Plan')
    compare_file = read_cmp_file()
    write_data(sheets_value, new_tar_file, compare_file)
    compare_data = make_cmp_data(sheets_value)
    save_cmp_file(compare_data)


if __name__ == '__main__':
    main()
