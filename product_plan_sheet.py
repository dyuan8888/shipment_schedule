#!ProductPlanSheet.py
"""Extract the data from Product Plan and write them to a new spreadsheet"""

from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Open the source file to read data from
def open_src_file(src_file, *sheets):
    wb = load_workbook(src_file, read_only=True, data_only=True)
    for sheet in sheets:
        ws = wb[sheet]
        if sheet == 'M-Plan':
            mo_list = extract_data(ws,rx=3,ry=8)
        else:
            etch_list = extract_data(ws,rx=5,ry=10)    
    sheets_value = mo_list + etch_list
    return sheets_value



# Extract data from the MFG product plan
def extract_data(ws,rx,ry):
    total_cell_values=[]
    now = datetime.datetime.now()
    future= now +datetime.timedelta(days=60)
    for i in range(2,90):
        cell_values=[]
        for j in range(rx, ry):
            cell_value = ws.cell(row=i, column=j).value
            cell_values.append(cell_value)
        if cell_values[1] != None and 'NSO' not in cell_values[1] and 'Upgr' not in cell_values[1] and now<cell_values[4]<future and 'R&D' not in cell_values[0]:
            cell_values = format_date(cell_values)
            total_cell_values.append(cell_values)
    return total_cell_values



# Format the date
def format_date(cell_values):
    cell_values[3] = datetime.datetime.strftime(cell_values[3],'%Y-%m-%d')
    cell_values[4] = datetime.datetime.strftime(cell_values[4],'%Y-%m-%d')
    return cell_values



# Write data in the target file
def write_data(sheets_value, tar_file):
    wb = Workbook()
    ws = wb.active
    row_title_style(wb,ws)
    for row_value in sheets_value:
        # TODO:Compare the data
        #row_value1 = compare_data(row_value, tar_file)
        ws.append(row_value)
    wb.save(tar_file)


# Compare new data with old data in the shipment schedule
def compare_data(row_value, tar_file):
    pass
            

# Style the first-row title
def row_title_style(wb, ws):
    title = ['Customer', 'Product Info', 'Project ID', 'Crate Date', 'Ship Date']
    ws.append(title)
    for m in range(1,6):
        ws.column_dimensions[get_column_letter(m)].width = 20
        ws.cell(row =1, column = m).font = Font(bold=True, size=14)


def main():
    #filename= input('Please enter the source file: ')
    src_file = r'c:\users\danielyuan\desktop\ProductPlanProject\Prod Plan 2019-W15 041219.xlsx'
    tar_file = r'c:\users\danielyuan\desktop\ProductPlanProject\Shipment Schedule.xlsx'
    sheets_value = open_src_file(src_file, 'M-Plan', 'E-Plan')
    #pprint.pprint(sheets_value)
    write_data(sheets_value, tar_file)
    



if __name__ == '__main__':
    main()
