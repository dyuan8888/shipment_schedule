# -*- coding: utf-8 -*-
"""
Created on Fri Apr 19 13:29:58 2019

@author: DanielYuan
"""

#!shipment schedule.py
"""Extract the data from Product Plan and write them to a new spreadsheet"""

from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import sys
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.colors import RED, BLUE


# Open the source file to read data from
def open_src_file(src_file, *sheets):
    wb = load_workbook(src_file, read_only=True, data_only=True)
    print('Start reading data from the source file...')
    for sheet in sheets:
        ws = wb[sheet]
        if sheet == 'M-Plan':
            mo_list = extract_data(ws,rx=3,ry=8)
        else:
            etch_list = extract_data(ws,rx=5,ry=10)    
    sheets_value = mo_list + etch_list
    return sheets_value


# Extract data from the MFG product plan
def extract_data(ws,rx,ry):
    total_cell_values=[]
    now = datetime.datetime.now()
    future= now +datetime.timedelta(days=60)
    for i in range(2,90):
        cell_values=[]
        for j in range(rx, ry):
            cell_value = ws.cell(row=i, column=j).value
            cell_values.append(cell_value)
        if cell_values[1] != None and 'NSO' not in cell_values[1] and 'Upgr' not in cell_values[1] and now<cell_values[4]<future and 'R&D' not in cell_values[0]:
            cell_values = format_date(cell_values)
            total_cell_values.append(cell_values)
    return total_cell_values


# Format the date
def format_date(cell_values):
    cell_values[3] = datetime.datetime.strftime(cell_values[3],'%Y-%m-%d')
    cell_values[4] = datetime.datetime.strftime(cell_values[4],'%Y-%m-%d')
    return cell_values


# Write data in the target file
def write_data(sheets_value, new_tar_file, cur_tar_file):
    wb = Workbook()
    ws = wb.active
    row_title_style(wb,ws)  # Call to generate column title
    ship_dict = get_cur_ship_data(cur_tar_file)
    print('Starting writing data to the target file...')
    num = 0    
    for row_value in sheets_value:
        row_value = compare_data(row_value, ship_dict)
        ws.append(row_value)
        num += 1
        if num < 2:
            print('%s row added...' %num)
        else:
            print('%s rows added...' %num)
        color_font(ws)
    wb.save(new_tar_file)
    print('\nA total of %d shipments have been successfully added!' %num)


# Color the specified fonts
def color_font(ws):
    for cell in list(ws.columns)[2]:
        if '*' in cell.value:
            cell.font = Font(color=BLUE)
        elif '!' in cell.value:
            cell.font = Font(color=RED)


# Compare new ship data with existing ship data
def compare_data(row_value, ship_dict):
    row_value[2] = str(row_value[2])
    if row_value[2] in ship_dict.keys():
        if row_value[3] != ship_dict[row_value[2]][0] or row_value[4] != ship_dict[row_value[2]][1]:
            row_value[2] = row_value[2] + '!'
    elif row_value[2] not in ship_dict.keys():
        row_value[2] = row_value[2] + '*'
    return row_value 


# Get the current shipment schedule
def get_cur_ship_data(cur_tar_file):
    try:
        wb = load_workbook(cur_tar_file)
    except FileNotFoundError:
        sys.exit("Sorry, the file '%s' that you just entered does not exist!" % cur_tar_file)
    else:
        ws = wb.active
        ship_dict = {}
        for x in range(2, ws.max_row+1):
            cell_value_3 =str(ws.cell(row=x, column=3).value).strip('*').strip('!')
            cell_value_4 = ws.cell(row=x, column=4).value
            cell_value_5 = ws.cell(row=x, column=5).value
            ship_dict[cell_value_3] = [cell_value_4, cell_value_5]
        return ship_dict


# Style the first-row title
def row_title_style(wb, ws):
    title = ['Customer', 'Product Info', 'Project ID', 'Crate Date', 'Ship Date']
    ws.append(title)
    for col in range(1,6):
        ws.column_dimensions[get_column_letter(col)].width = 20
        ws.cell(row =1, column = col).font = Font(bold=True, size=14)

def file_names():
    #filename1= input('Please enter the MFG product plan file: ')
    now = datetime.datetime.now()
    now_str = datetime.datetime.strftime(now, '%m%d%Y')
    src_file = 'c:\\users\\danielyuan\\desktop\\ProductPlanProject\\Prod Plan 2019-W15 041219.xlsx'
    new_tar_file = 'c:\\users\\danielyuan\\desktop\\ProductPlanProject\\Shipment Schedule_%s.xlsx' % now_str
    filename2 = input('Please enter the current shipment schedule file: ')
    cur_tar_file = 'c:\\users\\danielyuan\\desktop\\ProductPlanProject\\%s.xlsx' %filename2
    return src_file, new_tar_file, cur_tar_file




def main():
    src_file, new_tar_file, cur_tar_file = file_names()
    sheets_value = open_src_file(src_file, 'M-Plan', 'E-Plan')
    write_data(sheets_value, new_tar_file, cur_tar_file)


if __name__ == '__main__':
    main()
